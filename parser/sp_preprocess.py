# готовы справочники 1-21
# добавлен список из 6го справочника
## Создает справочник в удобном для pandas формате
    # dependences: find_number_page, read_text_documents

from os.path import exists, join
import pandas as pd
import re
import pickle as pkl
import pdfplumber
from src.app.checker.utils.text_utils import read_text_documents, find_number_page


def read_sp1(PATH, f_name):
    import openpyxl
    import pandas as pd
    import pickle as pkl
    from os.path import exists, join
    from openpyxl.styles.numbers import BUILTIN_FORMATS_MAX_SIZE, BUILTIN_FORMATS
    import re

    PKL_NAME = "sp1"  # имя файла для сохранения
    if exists(join(PATH, PKL_NAME+'.pkl')):
        print("Файл СП1 уже существует")
        return True

    # игнор ошибок в форматировании для openpyxl
    def monkey_patch_openpyxl():
        '''Openpyxl has a bug with workbooks that have wrong cell styling information.
        Monkey patch the library so it can handle these types of workbooks.'''
        from openpyxl.styles import stylesheet

        def _expand_named_style(self, named_style):
            """
            Bind format definitions for a named style from the associated style
            record
            """
            try:
                xf = self.cellStyleXfs[named_style.xfId]
                named_style.font = self.fonts[xf.fontId]
                named_style.fill = self.fills[xf.fillId]
                named_style.border = self.borders[xf.borderId]

                if xf.numFmtId < BUILTIN_FORMATS_MAX_SIZE:
                    formats = BUILTIN_FORMATS
                else:
                    formats = self.custom_formats
                if xf.numFmtId in formats:
                    named_style.number_format = formats[xf.numFmtId]
                if xf.alignment:
                    named_style.alignment = xf.alignment
                if xf.protection:
                    named_style.protection = xf.protection
            except:
                pass

        stylesheet.Stylesheet._expand_named_style = _expand_named_style

    # monkey_patch_openpyxl()

    ABS_PATH = join(PATH, f_name)
    print(ABS_PATH)
    data = pd.read_excel(ABS_PATH)
    data = data.iloc[:].dropna().reset_index(drop=True)
    data = data.rename(columns=data.iloc[0, :])
    data = data.drop([0, 1]).reset_index(drop=True)
    data.insert(1, "Архивный отдел", [''] * data.shape[0])

    workbook = openpyxl.load_workbook(ABS_PATH)  # Собственно - читаем сам файл
    sheets_list = workbook.sheetnames  # Получаем список всех листов в книге
    sheet_active = workbook[sheets_list[0]]  # Делаем активным самый первый лист в книге

    CELL_FILL = sheet_active['A14'].fill.start_color.index  # Получаем цвет ячейки
    SHIFT = 14  # смещение до первой ячейки таблицы
    idx = []
    # выбрали все строки нужного цвета
    for i in range(data.shape[0]):
        if sheet_active['A' + str(SHIFT + i)].fill.start_color.index == CELL_FILL:
            idx.append(i)
    idx_reserve = idx[:]
    idx.append(data.shape[0] - 1)
    idx = idx[::-1]
    n_idx = idx.pop()

    # почистили таблицу
    header = data.iloc[0, 0]
    n_idx = idx.pop()
    for row in range(data.shape[0]):
        if row < n_idx:
            data.at[row, "Архивный отдел"] = header
        #         if n_idx > 1271:
        #             print(header)
        elif len(idx) > 0:
            n_idx = idx.pop()
            header = data.iloc[n_idx, 0]
            data.at[row, "Архивный отдел"] = header
    data.at[row, "Архивный отдел"] = header
    for row in idx_reserve:
        data = data.drop([row])
    data = data.reset_index(drop=True)

    with open(f'{join(PATH,PKL_NAME)}.pkl', 'wb') as file:
        pkl.dump(data, file)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]


#
def read_sp2(PATH, f_name):
    import pandas as pd
    import pickle as pkl
    from os.path import exists, join
    PKL_NAME = "sp2"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП2 уже существует")
        return True
    # чтение СП2

    dF = pd.read_excel(join(PATH, f_name))
    data = dF.iloc[:,1]
    data = data[2:]
    data = data.reset_index(drop=True)

    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]


def read_sp3(PATH, f_name):
    import pandas as pd
    import pickle as pkl
    from os.path import exists, join
    PKL_NAME = "sp3"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП3 уже существует")
        return True
    # чтение СП3
    ABS_PATH = join(PATH, f_name)
    data = pd.read_excel(ABS_PATH)
    data = data.iloc[0:, 1:]
    data = data.rename(columns=data.iloc[1, :])
    data = data.iloc[2:, :]
    data = data.reset_index(drop=True)
    data.insert(2, "Архивное подразделение", [''] * data.shape[0])
    data.insert(3, "Наименование подразделения", [''] * data.shape[0])

    # убрать заголовки и внести их в таблицу
    arch_val = ""
    dep_val = ""
    drops = []
    for row in range(data.shape[0]):
        d = data.iloc[row, :][0]
        if "Архив:" in d:
            arch_val = d[d.find(':') + 1:]
            drops.append(row)
        if "Наименование:" in d:
            dep_val = d[d.find(':') + 1:]
            drops.append(row)
        data.at[row, "Архивное подразделение"] = arch_val
        data.at[row, "Наименование подразделения"] = dep_val
    data = data.reset_index(drop=True)
    drops.append(data.shape[0] - 1)
    drops = list(set(drops))
    for row in drops:
        data = data.drop([row])
    data = data.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]

def read_sp4(PATH, f_name):

    def get_table_documents_sp4(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = 0
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        return table

    def proc_sp4(PATH, PKL_NAME):  # предобработать pkl-файл
        if exists(join(PATH, f'{PKL_NAME}_proc.pkl')):
            print("Файл СП4 с предобработкой уже существует")
            return True
        with open(join(PATH, 'sp4.pkl'), 'rb') as f:
            table = pkl.load(f)
        # заголовок
        cols = [str(x) for x in table.iloc[0, :]]
        for i in range(len(cols)):
            if "\nие" in cols[i]:
                cols[i] = cols[i].replace("\nие", "ие")
            if "\nество" in cols[i]:
                cols[i] = cols[i].replace("\nество", "ество")
            if cols[i][-5:] == "едини":
                cols[i] = cols[i] + "ц"
        table.columns = [val.replace("\n", " ") for val in cols]

        drops = [0]
        for row in range(table.shape[0]):
            val = str(table.iloc[row, 0])
            if ":" in val:
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        for row in range(table.shape[0]):
            for cell in range(table.shape[1]):
                if "\n" in table.iloc[row, cell]:
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\n", " ")
        return table

    PKL_NAME = "sp4"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП4 уже существует")
    else:
        ABS_PATH = join(PATH, f_name)
        table = get_table_documents_sp4(ABS_PATH)
        data = table.reset_index(drop=True)
        with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
            pkl.dump(data, f)
            print(f"Файл справочника {PKL_NAME}.pkl создан")
    if exists(join(PATH, f'{PKL_NAME}_proc.pkl')):
        print("Файл СП4 с предобработкой уже существует")
    else:
        PKL_NAME = "sp4"
        table = proc_sp4(PATH, PKL_NAME)
        data = table.reset_index(drop=True)
        with open(f'{join(PATH, PKL_NAME)}_proc.pkl', 'wb') as f:
            pkl.dump(data, f)
            print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}_proc.pkl')]
    

def read_sp5(PATH, f_name):
    from src.app.checker.utils.text_utils import find_number_page, read_text_documents
    import pdfplumber
    def add_header_cols_sp5(pages):
        import re
        from string import digits
        col_header = ""
        col_sub_header = ""
        col_sub2_header = ""
        pages.insert(4, "Раздел", [''] * pages.shape[0])
        pages.insert(5, "Подраздел", [''] * pages.shape[0])
        pages.insert(6, "Под-подраздел", [''] * pages.shape[0])

        drops = []
        pages = pages.reset_index(drop=True)

        for row in range(pages.shape[0]):
            if pages.iloc[row][0] is not None:
                for i, el in enumerate(pages.iloc[row][0]):  # убрать избыточные переносы
                    if el == '\n':
                        if len(pages.iloc[row][0]) > i + 1 and pages.iloc[row][0][i + 1] not in digits:
                            val = pages.iloc[row][0][:]
                            val = val[:i] + ' ' + val[i + 1:]
                            pages.loc[row, 0] = val

                if re.search(r'^\d+\.\s*[А-Я]+', pages.iloc[row][0]):  # заголовок 1 уровня
                    drops.append(row)
                    if '\n' in pages.iloc[row][0]:
                        col_header = pages.iloc[row][0].split('\n')[0]
                    else:
                        col_header = pages.iloc[row][0]
                if re.search(r'\n\d+\.\d+\.\s*[А-Я]+', pages.iloc[row][0]):  # заголовок 2 уровня
                    col_sub_header = pages.iloc[row][0].split('\n')[1]
                if re.search(r'^\d+\.\d+\.\s*[А-Я]+', pages.iloc[row][0]):
                    drops.append(row)
                    col_sub_header = pages.iloc[row][0].split('\n')[0]
                if len(col_header) > 0 and len(col_sub_header) > 0 and col_header[0] != col_sub_header[
                    0]:  # если нет заголовка 2го уровня
                    col_sub_header = ""
                if re.search(r'^\d+\.\d+\.\d+\.\s*[А-Я]+', pages.iloc[row][0]):  # заголовок 3го уровня
                    drops.append(row)
                    col_sub2_header = pages.iloc[row][0][
                                      re.search(r'^\d+\.\d+\.\d+\.\s*[А-Я]+', pages.iloc[row][0]).start():]
                if re.search(r'\n\d+\.\d+\.\d+\.\s*[А-Я]+', pages.iloc[row][0]):  # заголовок 3го уровня
                    col_sub2_header = pages.iloc[row][0][
                                      re.search(r'\n\d+\.\d+\.\d+\.\s*[А-Я]+', pages.iloc[row][0]).start() + 1:]
                if len(col_sub_header) > 0 and len(col_sub2_header) > 0:
                    left = col_sub_header.split('.')
                    right = col_sub2_header.split('.')
                    if left[1] != right[1]:  # если нет заголовка 3го уровня
                        col_sub2_header = ""

            pages.at[row, "Подраздел"] = col_sub_header
            pages.at[row, "Под-подраздел"] = col_sub2_header
            pages.at[row, "Раздел"] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def remove_header_lines_sp5(table):  # Убрать заголовочные строки
        drops = []
        for i, line in table.iterrows():
            if all(line.values[:4] == ['1', '2', '3', '4']):
                drops.append(i)
                break
            drops.append(i)
        table.drop(table.index[drops], inplace=True)
        table = table.reset_index(drop=True).fillna('').astype(str)
        return table

    def merge_border_cells_sp5(table):  # объединить смежные ячейки на разных страницах
        indx = []
        for row in range(table.shape[0]):
            if table.iloc[row, 0] == '':
                if table.iloc[row - 1, 0]:
                    indx.append(row)
        cols = list(table.columns)[1:]
        for row in indx:
            for col in cols:
                table.at[row - 1, col] = table.iloc[row - 1][col] + '\n' + table.iloc[row][col]
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True).fillna('').astype(str)
        return table

    def detect_list_sp5(table):  # найти список из букв
        indx = []
        for row in range(table.shape[0]):  # нашли номера строк, где есть списки
            val = table.iloc[row, 1]
            if val.count(")") > val.count("(") and 'а)' in val:
                indx.append(row)
        for i, row in enumerate(indx):  # нашли позицию пунктов списка
            matches = []
            for match in re.finditer(r'\n[а-я]\)\s', table.iloc[row, 1]):
                matches.append(match.span()[0])
            indx[i] = [row] + matches
        for j, row in enumerate(indx):  # получили элементы списка
            lst = [table.iloc[row[0], 1][:row[1] + 1]]
            for i in range(1, len(row) - 1):
                lst.append(table.iloc[row[0], 1][row[i] + 1:row[i + 1]])
            lst.append(table.iloc[row[0], 1][row[-1] + 1:])
            indx[j] = [indx[j][0]] + lst
        for j, row in enumerate(indx):  # элементы 2-го столбца, как отдельные записи
            for i in range(2, len(row)):
                row[i] = row[i][:2] + row[1][:-2] + " " + row[i][2:]
            del row[1]
            indx[j] = row[:]
        return indx

    def insert_indx_sp5(table, indx):  # создать дубликаты строк
        new_table = pd.DataFrame(columns=table.columns)
        indx_rows = [i[0] for i in indx]
        val_rows = [i[1:] for i in indx]
        i = 0
        j_indx = 0
        for row in range(table.shape[0]):
            if row not in indx_rows:
                new_table.loc[i] = table.iloc[row]
                i += 1
            else:
                inserted = val_rows[j_indx]
                j_indx += 1
                for val in inserted:
                    new_table.loc[i] = table.iloc[row]
                    new_table.loc[i, new_table.columns[1]] = val
                    i += 1

        return new_table

    def get_table_documents_sp5(path):
        table_settings = {
            "explicit_horizontal_lines": [715],  # добавить линию в конце страницы
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'II. '.lower())
        number_end = find_number_page(text, '<1> Срок хранения'.lower())
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])

        table = add_header_cols_sp5(table)  # добавить подразделы в таблицу
        table.rename(columns={0: "Номер статьи", 1: "Вид документа", 2: "Срок хранения документа", 3: "Примечания"},
                     inplace=True)
        table = remove_header_lines_sp5(table)
        table = merge_border_cells_sp5(table)
        indx = detect_list_sp5(table)
        table = insert_indx_sp5(table, indx)
        return table

    PKL_NAME = "sp5"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП5 уже существует")
        return True, [join(PATH, f'{PKL_NAME}.pkl')]
    # чтение СП5
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp5(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]

def read_sp5_list(PATH, f_name):
    def find_numbers(cell):
        if re.search(r'\d+', cell):
            return True
        return False

    def get_list_sp5(path):
        text, full_text, page_blocks = read_text_documents(path)

        number_begin = find_number_page(text, 'УКАЗАТЕЛЬ ВИДОВ ДОКУМЕНТОВ'.lower())
        number_end = len(text) - 1

        nec_text = []
        for i in range(number_begin, number_end + 1):
            if 'УКАЗАТЕЛЬ ВИДОВ ДОКУМЕНТОВ' not in text[i]:
                nec_text.append(text[i])
            else:
                start = re.search('УКАЗАТЕЛЬ ВИДОВ ДОКУМЕНТОВ', text[i]).regs[0][0]
                nec_text.append(text[i][start:])

        chapter = ""
        columns = ["Название", "Расположение", "Раздел"]
        table = pd.DataFrame(columns=columns)
        k = 0
        val = ""
        num = ""
        nec_text[0] = "\n".join(nec_text[0].split("\n")[1:])
        drops = []
        for i in range(len(nec_text)):
            line = nec_text[i]
            line_split = line.split('\n')

            for j, el in enumerate(line_split):
                if find_numbers(el):
                    num = el
                if re.search(r'^[А-Я]+$', el):
                    chapter = el
                    if line_split[j + 1] != "\xa0" and find_numbers(line_split[j + 1]):
                        table.loc[len(table)] = {table.columns[0]: el, table.columns[1]: line_split[j + 1],
                                                 table.columns[2]: chapter}
                        drops.append(len(table))
                else:
                    if not find_numbers(el):
                        val = val + el
                    if find_numbers(el):
                        num = el
                        if "\xa0" in val:
                            val = val.replace("\xa0", "")
                        table.loc[len(table)] = {table.columns[0]: val, table.columns[1]: num,
                                                 table.columns[2]: chapter}
                        val = ""
                        num = ""
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        for i in range(len(table)):
            if table.iloc[i, 0] == "":
                vals = table.iloc[i, 1]
                num_start = re.search(r'\d+', vals).start()
                table.iloc[i, 0] = vals[:num_start]
                table.iloc[i, 1] = vals[num_start:]
        drops = []
        for i in range(len(table)):
            if table.iloc[i, 0] != "":
                base_i = i
            else:
                table.iloc[base_i, 1] = table.iloc[base_i, 1] + table.iloc[i, 1]
                drops.append(i)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table

    PKL_NAME = "sp5"
    if exists(join(PATH, f'{PKL_NAME}_list.pkl')):
        print(f"Файл СП5.список уже существует")
        # return True
    # чтение СП5
    ABS_PATH = join(PATH, f_name)

    table = get_list_sp5(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}_list.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}_list.pkl создан")
    return True

def read_sp6(PATH, f_name):
    from src.app.checker.utils.text_utils import find_number_page, read_text_documents

    def add_header_cols_sp6(pages):
        import re
        indx = []
        for row in range(pages.shape[0] - 1):
            if pages.iloc[row, 0] is not None and len(pages.iloc[row, 1:-1]) == list(pages.iloc[row, 1:-1]).count(None):
                if pages.iloc[row + 1, 0] is not None and len(pages.iloc[row, 1:-1]) == list(
                        pages.iloc[row, 1:-1]).count(
                        None) and list(pages.iloc[row, 1:-1]).count(None) > 0:
                    if re.search(r'^\d{,4}\.\d{,4}\.', pages.iloc[row, 0]) is not None and re.search(r'^\d{,4}\.',
                                                                                                     pages.iloc[
                                                                                                         row + 1, 0]) is None:
                        pages.iloc[row, 0] = pages.iloc[row, 0] + " " + pages.iloc[row + 1, 0]
                        # indx.append(row)
                        indx.append(row + 1)
        pages = pages.drop(indx)
        pages = pages.reset_index(drop=True)

        col_sub_header = ""
        pages.insert(pages.shape[1], "Подраздел", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)

        for row in range(pages.shape[0]):
            header_row_vals = [pages.iloc[row, i] for i in range(1, len(pages.iloc[row]) - 1)]
            if pages.iloc[row, 0] is not None and header_row_vals.count(None) == len(header_row_vals):
                if re.search(r'^\d+\.\d+\.\s*[А-Я]+', pages.iloc[row, 0]):
                    drops.append(row)
                    col_sub_header = pages.iloc[row, 0]
            pages.at[row, "Подраздел"] = col_sub_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        # Объединить строки заголовка раздела

        return pages

    def merge_cells_sp6(table):  # объединить смежные ячейки
        indx = []
        for row in range(table.shape[0]):
            table_row_vals = [table.iloc[row, i] for i in range(table.shape[1])]
            table_row_vals = [table_row_vals[0]] + table_row_vals[2: table.shape[1] - 3]
            if table_row_vals.count('') == len(table_row_vals):
                indx.append(row)
        cols = list(table.columns)[1:-1]
        for row in indx:
            row_base = row - 1
            while row + 1 in indx:
                for col in cols:
                    table.at[row_base, col] = str(table.iloc[row_base][col]) + '\n' + str(table.iloc[row][col])
                row += 1
            for col in cols:
                table.at[row_base, col] = str(table.iloc[row_base][col]) + '\n' + str(table.iloc[row][col])
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True)
        # объединить строчки для одной буквы
        indx = []
        for row in range(table.shape[0]):  # нашли номера строк, где есть списки
            if table.iloc[row, 0] is None and table.iloc[row + 1, 0] == "":
                for col in range(1, table.shape[1] - 1):
                    row_0, row_1 = table.iloc[row, col], table.iloc[row + 1, col]
                    table.at[row, table.columns[col]] = str(row_0) + '\n' + str(row_1)
                    indx.append(row + 1)
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True)
        # объединить для букв
        indx = []
        for row in range(table.shape[0] - 1):  # нашли номера строк, где есть списки
            row_0 = [table.iloc[row, 0], table.iloc[row, 1]]
            row_next = [table.iloc[row + 1, 0], table.iloc[row + 1, 1]]
            if row_0[0] is None and row_next[0] is None:
                if re.search(r'^[а-я]\)\s', str(row_0[1])) and re.search(r'^[а-я]\)\s', str(row_next[1])) is None:
                    for col in range(1, table.shape[1] - 1):
                        if table.iloc[row + 1, col] is not None:
                            table.at[row, table.columns[col]] = str(table.iloc[row, col]) + '\n' + str(
                                table.iloc[row + 1, col])
                    indx.append(row + 1)
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True)
        # исправление дефекта таблицы 1го раздела (смещение столбцов)
        indx = []
        for row in range(table.shape[0]):  # нашли номера строк
            if table.iloc[row, 0] is not None and re.search(r'^\d{,4}\.', table.iloc[row, 0]) is None:
                if len(table.iloc[row, 0]) > 0:
                    table.loc[row] = [None] + list(table.iloc[row])[:-2] + [table.iloc[row, -1]]
        for row in range(table.shape[0] - 1):  # нашли номера строк
            if table.iloc[row, 0] is None and table.iloc[row + 1, 0] is None:
                if re.search(r'^[а-я]\)\s', str(table.iloc[row, 1])) and re.search(r'^[а-я]\)\s',
                                                                                   str(table.iloc[row + 1, 1])) is None:
                    for col in range(1, table.shape[1] - 1):
                        if table.iloc[row + 1, col] is not None:
                            table.at[row, table.columns[col]] = str(table.iloc[row, col]) + '\n' + str(
                                table.iloc[row + 1, col])
                    indx.append(row + 1)
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True)
        return table

    def detect_list_sp6(table):  # найти список из букв и объединить строки
        indx = []
        indx_bases = []
        for row in range(table.shape[0]):  # еще проблемка с буквами в списках
            if table.iloc[row, 0] == "" and len(str(table.iloc[row, 1])) > 2:
                if str(table.iloc[row, 1])[1] == ")":
                    table.iloc[row, 0] = None
        for row in range(table.shape[0]):
            val = table.iloc[row, 1]
            if row + 1 < table.shape[0] and (table.iloc[row, 0] is not None) and table.iloc[row + 1, 0] is None:
                indx_bases.append(row)
            if val is not None and len(val) > 1 and val[1] == ")":
                indx.append(row)
        new_table = pd.DataFrame(columns=table.columns)
        i = 0
        for row in range(table.shape[0]):
            if row not in indx and row not in indx_bases:
                new_table.loc[i] = table.iloc[row]
                i += 1
            else:
                if row in indx and row not in indx_bases:
                    ind = [x for x in indx_bases if x < row][-1]
                    for col in range(table.shape[1]):
                        if col == 0:
                            new_table.loc[i, table.columns[col]] = table.iloc[ind, col]
                        if col == 1:
                            new_table.loc[i, table.columns[col]] = str(table.iloc[row, col])[:2] + " " + table.iloc[
                                ind, col] + ":" + str(table.iloc[row, col])[3:]
                        if col not in [0, 1]:
                            new_table.loc[i, table.columns[col]] = table.iloc[row, col]
                    i += 1
        return new_table

    def repair_problems(table):
        # строка, в которой оказалось в одной ячейке заголовок и элемент списка
        indx = []
        for row in range(1, table.shape[0] - 1):
            if table.iloc[row, 0] is None and table.iloc[row, 1] is None:
                if table.iloc[row, 2] is not None and table.iloc[row, 3] is not None:
                    if len(table.iloc[row, 2]) > 2 and len(table.iloc[row, 3]) > 2:
                        if str(table.iloc[row + 1, 1])[1] == ")":  # список
                            val = table.iloc[row + 1, 1].split(":")
                            table.iloc[row + 1, 1] = val[0] + ":" + val[-1]
                            table.iloc[row, 1] = (val[1][:3] + val[0][3:] + val[1][3:]).strip()
                            table.iloc[row, 0] = table.iloc[row - 1, 0]
        # строка, в которой разрыв между страницами в ячейке
        for row in range(1, table.shape[0] - 1):
            if table.iloc[row, 0] is None and all(table.iloc[row, 1:3]) is not None:
                if len(table.iloc[row, 1]) > 2 and len(table.iloc[row, 2]) == 0 and len(table.iloc[row, 3]) == 0:
                    table.iloc[row - 1, 1] = table.iloc[row - 1, 1] + table.iloc[row, 1]
                    indx.append(row)
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True)

        # многострочные пункты
        indx = []
        for row in range(1, table.shape[0] - 1):
            if table.iloc[row, 0] == "" and table.iloc[row, 1] is not None and re.search(r'^\d{,4}\.',
                                                                                         table.iloc[row - 1, 0]):
                if len(table.iloc[row, 1]) > 0:
                    for col in range(table.shape[1] - 1):
                        table.iloc[row - 1, col] = str(table.iloc[row - 1, col]) + ' ' + str(table.iloc[row, col])
                    indx.append(row)
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True)

        # многострочные ячейки в 3м столбце
        indx = []
        for row in range(1, table.shape[0] - 1):
            if table.iloc[row, 0] == "" and table.iloc[row, 1] == "" and table.iloc[row, 2] is not None:
                for col in range(table.shape[1] - 1):
                    table.iloc[row - 1, col] = str(table.iloc[row - 1, col]) + " " + str(table.iloc[row, col])
                indx.append(row)
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True)
        return table

    def make_header_sp6(table):  # сделать заголовок для таблицы и дропнуть верхнюю и нижнюю часть таблицы
        # дропнуть верхнюю часть таблицы, не относящуюся к разделу
        for row in range(table.shape[0]):
            vals = list(table.iloc[row])
            header_lines_num = table.shape[0]
            if vals[0] in ['N п/п', 'N\nп/п']:
                header_lines_num = row
                break
        table = table.drop(range(0, header_lines_num))
        table = table.reset_index(drop=True)
        for row in range(table.shape[0]):
            vals = list(table.iloc[row])
            if vals[:3] == ['1', '2', '3']:
                header_lines_num = row
                break
        for row in range(header_lines_num):  # продублировать заголовки в строке
            vals = table.iloc[row]
            for i, col in enumerate(vals):
                if col is None and i != 0:
                    vals[i] = vals[i - 1]
            table.iloc[row] = vals

        for col in table.columns:  # объединить строки заголовка
            val = ""
            for row in range(header_lines_num):
                val = val + " " + str(table.iloc[row][col])
            table.iloc[0, col] = val.replace('\n', ' ')
        table.columns = table.iloc[0]
        table = table.drop(range(header_lines_num + 1))
        table = table.reset_index(drop=True)
        # дропнуть нижнюю часть таблицы, не относящуюся к разделу
        for row in range(table.shape[0]):
            vals = list(table.iloc[row])
            header_lines_num = table.shape[0]
            if vals[0] in ['N п/п', 'N\nп/п'] or vals[1] in ['N п/п', 'N\nп/п']:
                header_lines_num = row
                break
        table = table.drop(range(header_lines_num, table.shape[0]))
        table = table.reset_index(drop=True)
        # дропнуть пустой столбец
        indx = []
        for col in range(table.shape[1]):
            vals = list(table.iloc[:, col])
            if vals.count(vals[0]) == len(vals):
                indx.append(col)
        table = table.drop(table.columns[indx], axis=1)
        return table

    def clear_cells(table):  # очистка таблицы, ужасные переносы, вариантов бесконечно, лучше стемминг провести
        header = list(table.columns)
        for ind in range(len(header)):
            val = header[ind]
            if "None" in val:
                header[ind] = val.replace("None", "")
        table.columns = header
        for row in range(table.shape[0]):
            for cell in range(table.shape[1]):
                if table.iloc[row, cell] is None:  # заменить None на "-"
                    table.iloc[row, cell] = "-"
                if "None" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("None", "")
                if "\n\n" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\n\n", " ")
                table.iloc[row, cell] = str(table.iloc[row, cell]).strip()
                if "::" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("::", ": ")
                if ",\n" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace(",\n", ", ")
                if "\n," in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\n,", ",")
                if "-\n" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("-\n", "-")
                if "\nы" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nы", "ы")
                if "\n(" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\n(", " (")
                if "\n)" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\n)", ")")
                if str(table.iloc[row, cell])[-2:] == "\nо":
                    table.iloc[row, cell] = table.iloc[row, cell][:-2] + "о"
                if str(table.iloc[row, cell])[-2:] == "\nи":
                    table.iloc[row, cell] = table.iloc[row, cell][:-2] + "и"
                if "\nх" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nх", "х")
                if "\nй" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nй", "й")
                if "\nой" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nой", "ой")
                if "\nной" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nной", "ной")
                if "\nий" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nий", "ий")
                if "\nие" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nие", "ие")
                if "\nии" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nии", "ии")
                if "\nые" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nые", "ые")
                if "\nские" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nские", "ские")
                if "\nтва" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nтва", "тва")
                if "\nные" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nные", "ные")
                if "\nных" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nных", "ных")
                if "\nьных" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nьных", "ьных")
                if "\nсти" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nсти", "сти")
                if "\nний" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nний", "ний")
                if "\nв," in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nв,", "в,")
                if "\nю" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nю", "ю")
                if "\nнгу" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nнгу", "нгу")
                if "\nнге" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nнге", "нге")
                if "\nого" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nого", "ого")
                if "ы\nе" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("ы\nе", "ые")
                if "\nя," in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nя,", "я,")
                if "\nвенных" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nвенных", "венных")
                if "\nческих" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nческих", "ческих")
                if "\nорских" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nорских", "орских")
                if "\nонной" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nонной", "онной")
                if "\nонных" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nонных", "онных")
                if "\nьно" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nьно", "ьно")
                if "\nния" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nния", "ния")
                if "\nвного" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nвного", "вного")
                if "м\nмы" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("м\nмы", "ммы")
                if "\nмм" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nмм", "мм")
                if "к\nих" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("к\nих", "ких")
                if "к\nов" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("к\nов", "ков")
                if "\nовых" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nовых", "овых")
                if "\nо-" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nо-", "о-")
                if "\nаты" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nаты", "аты")
                if "\nение" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nение", "ение")
                if "\nанные" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nанные", "анные")
                if "\nеские" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nеские", "еские")
                if "\nнской" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nнской", "нской")
                if "\nтные" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nтные", "тные")
                if "\nска" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nска", "ска")
                if "\nение" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nение", "ение")
                if "\nения" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nения", "ения")
                if "\nнных" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nнных", "нных")
                if "\nских" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nских", "ских")
                if "\nвенные" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nвенные", "венные")
                if "\nации" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nации", "ации")
                if "н\nты" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("н\nты", "нты")
                if "\nнты" in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("\nнты", "нты")
                if "с\nка " in str(table.iloc[row, cell]):
                    table.iloc[row, cell] = str(table.iloc[row, cell]).replace("с\nка ", "ска ")
        return table

    def get_table_documents_sp6(path, chapter):

        table_settings = {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)
        if chapter != 12:
            number_begin = find_number_page(text, f'Раздел {chapter}.'.lower())
            number_end = find_number_page(text, f'Раздел {chapter + 1}.'.lower())
        else:
            number_begin = find_number_page(text, f'Раздел {chapter}.'.lower())
            number_end = find_number_page(text, 'УКАЗАТЕЛЬ ВИДОВ ДОКУМЕНТОВ'.lower())

        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        v_lines = [pdf_res.pages[i].bbox[1], pdf_res.pages[i].bbox[3]]
        h_lines = [pdf_res.pages[i].bbox[0], pdf_res.pages[i].bbox[2]]
        table_settings = {
            "explicit_vertical_lines": v_lines,  # добавить линию в конце страницы
            "explicit_horizontal_lines": h_lines,  # добавить линию в конце страницы
        }
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            tables = pdf_res.pages[i].find_tables(table_settings)
            for j in range(len(tables)):
                table = pd.concat([table, pd.DataFrame(tables[j].extract())])
        table = table.reset_index(drop=True)

        table = make_header_sp6(table)  # сделать заголовок
        table = add_header_cols_sp6(table)  # добавить подразделы в таблицу
        table = merge_cells_sp6(table)
        table = detect_list_sp6(table)
        table = repair_problems(table)  # починка проблемных ячеек
        table = clear_cells(table)
        return table

    PKL_NAME = "sp6"
    CHAPTERS = 12
    for chapter in range(1, CHAPTERS + 1):
        if exists(join(PATH, f'{PKL_NAME}_{chapter}.pkl')):
            print(f"Файл СП6.{chapter} уже существует")
            return True
    # чтение СП6
    fpaths = []
    ABS_PATH = join(PATH, f_name)
    for chapter in range(1, CHAPTERS + 1):
        table = get_table_documents_sp6(ABS_PATH, chapter)
        data = table.reset_index(drop=True)
        with open(f'{join(PATH, PKL_NAME)}_{chapter}.pkl', 'wb') as f:
            pkl.dump(data, f)
            print(f"Файл справочника {PKL_NAME}_{chapter}.pkl создан")
            fpaths.append(join(PATH, f'{PKL_NAME}_{chapter}.pkl'))
    return True, fpaths


def read_sp6_list(PATH, f_name):
    def find_numbers(cell):
        if re.search(r'\d+', cell):
            return True
        return False

    def get_list_sp6(path):
        text, full_text, page_blocks = read_text_documents(path)

        number_begin = find_number_page(text, 'УКАЗАТЕЛЬ ВИДОВ ДОКУМЕНТОВ'.lower())
        number_end = len(text) - 1

        nec_text = []
        for i in range(number_begin, number_end + 1):
            if 'УКАЗАТЕЛЬ ВИДОВ ДОКУМЕНТОВ' not in text[i]:
                nec_text.append(text[i])
            else:
                start = re.search('УКАЗАТЕЛЬ ВИДОВ ДОКУМЕНТОВ', text[i]).regs[0][0]
                nec_text.append(text[i][start:])

        chapter = ""
        line = nec_text[0]
        line_split = line.split('\n')
        columns = line_split[1:4]
        columns[1] = columns[1] + " " + columns[-1]
        columns[-1] = "Раздел"

        table = pd.DataFrame(columns=columns)
        k = 0
        val = ""
        nec_text[0] = "\n".join(line_split[4:])
        for i in range(len(nec_text)):
            line = nec_text[i]
            line_split = line.split('\n')
            for j, el in enumerate(line_split):

                if el.upper() == el and ":" in el:
                    chapter = el[:-1]
                else:
                    if not find_numbers(el):
                        val = val + el
                    if find_numbers(el):
                        num = el
                        if "\xa0" in val:
                            val = val.replace("\xa0", "")
                        table.loc[len(table)] = {table.columns[0]: val, table.columns[1]: num,
                                                 table.columns[2]: chapter}
                        val = ""
                        num = ""
        for i in range(len(table)):
            if table.iloc[i, 0] == "":
                vals = table.iloc[i, 1]
                num_start = re.search(r'\d+', vals).start()
                table.iloc[i, 0] = vals[:num_start]
                table.iloc[i, 1] = vals[num_start:]
        return table

    PKL_NAME = "sp6"
    if exists(join(PATH, f'{PKL_NAME}_list.pkl')):
        print(f"Файл СП6.список уже существует")
        return True
    # чтение СП6
    ABS_PATH = join(PATH, f_name)

    table = get_list_sp6(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}_list.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}_list.pkl создан")
    return True
    

def read_sp7(PATH, f_name):
    from src.app.checker.utils.text_utils import find_number_page, read_text_documents
    def add_header_cols_sp7(pages):
        import re
        from string import digits
        col_header = ""
        pages.insert(4, "Подразделение", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)

        for row in range(pages.shape[0]):
            if pages.iloc[row, 0] is not None:
                for i, el in enumerate(pages.iloc[row, 0]):  # убрать избыточные переносы
                    if el == '\n':
                        if len(pages.iloc[row][0]) > i + 1 and pages.iloc[row, 0][i + 1] not in digits:
                            val = pages.iloc[row, 0][:]
                            val = val[:i] + ' ' + val[i + 1:]
                            pages.loc[row, 0] = val

                if list(pages.iloc[row]).count(None) > 2:  # заголовок 1 уровня
                    drops.append(row)
                    if '\n' in pages.iloc[row, 0]:
                        col_header = pages.iloc[row, 0].split('\n')[0]
                    else:
                        col_header = pages.iloc[row, 0]
            pages.at[row, "Подразделение"] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def merge_border_cells_sp7(table):  # объединить смежные ячейки на разных страницах
        indx = []
        for row in range(table.shape[0]):
            if table.iloc[row, 0] == '':
                if table.iloc[row - 1, 0]:
                    indx.append(row)
        cols = list(table.columns)[1:]
        for row in indx:
            for col in cols:
                table.at[row - 1, col] = table.iloc[row - 1][col] + '\n' + table.iloc[row][col]
        table.drop(table.index[indx], inplace=True)
        table = table.reset_index(drop=True).fillna('').astype(str)
        return table

    def get_table_documents_sp7_1(path):
        table_settings = {
            # "explicit_horizontal_lines": [720],  # добавить линию в конце страницы
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'опись №5'.lower())
        number_end = find_number_page(text, 'опись №6'.lower()) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = table.iloc[0]
        table = table.drop([0])
        table = table.reset_index(drop=True)
        table = add_header_cols_sp7(table)  # добавить подразделы в таблицу
        return table

    def get_table_documents_sp7_2(path):
        table_settings = {
            "explicit_horizontal_lines": [720],  # добавить линию в конце страницы
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'опись №6'.lower())
        number_end = len(pdf_res.pages)-1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = table.iloc[0]
        table = merge_border_cells_sp7(table)
        table = table.reset_index(drop=True)
        return table

    PKL_NAME = "sp7"
    if exists(join(PATH, f'{PKL_NAME}_1.pkl')) and exists(join(PATH, f'{PKL_NAME}_2.pkl')):
        print("Файлы СП7 уже существуют")
        return True, [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]

    # чтение СП7
    ABS_PATH = join(PATH, f_name)
    table_1 = get_table_documents_sp7_1(ABS_PATH)
    data_1 = table_1.reset_index(drop=True)
    table_2 = get_table_documents_sp7_2(ABS_PATH)
    data_2 = table_2.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}_1.pkl', 'wb') as f:
        pkl.dump(data_1, f)
        print(f"Файл справочника {PKL_NAME}_1.pkl создан")
    with open(f'{join(PATH, PKL_NAME)}_2.pkl', 'wb') as f:
        pkl.dump(data_2, f)
        print(f"Файл справочника {PKL_NAME}_2.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]


def read_sp8(PATH, f_name):
   
    def add_header_cols_sp8(pages):
        col_header = ""
        pages.insert(4, "Раздел", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)
        for row in range(pages.shape[0]):
            if str(pages.iloc[row, 0]).upper() == str(pages.iloc[row, 0]) and len(pages.iloc[row, 0]) > 2:  # заголовок 1 уровня
                drops.append(row)
                col_header = pages.iloc[row, 0]
            pages.loc[row, pages.columns[4]] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def get_table_documents_sp8(path):
        import re
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = add_header_cols_sp8(table)
        table.drop(0, axis=0, inplace=True)
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            if not (table.iloc[row, 0] == ""):
                lst_indx = row
            else:
                for i in range(len(table.columns)):
                    table.loc[lst_indx, table.columns[i]] = (str(table.loc[lst_indx, table.columns[i]]) +
                                                             str(table.loc[row, table.columns[i]]))
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        for row in range(table.shape[0]):
            if re.search("\n-", " ".join(list(table.iloc[row]))) is None:
                table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            else:
                table_row = table.iloc[row]
                for i, cell in enumerate(table_row):
                    if re.search("\n-", cell) is None:
                        table_row[i] = cell.replace('\n', ' ')
        return table


def read_sp9(PATH, f_name):
    
    def add_header_cols_sp9(pages):
        col_header = ""
        pages.insert(4, "Раздел", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)
        for row in range(pages.shape[0]):
            if str(pages.iloc[row, 0]).upper() == str(pages.iloc[row, 0]) and len(pages.iloc[row, 0]) > 2:  # заголовок 1 уровня
                drops.append(row)
                col_header = pages.iloc[row, 0]
            pages.loc[row, pages.columns[4]] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def get_table_documents_sp9(path):
        import re
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = add_header_cols_sp9(table)
        table.drop(0, axis=0, inplace=True)
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            if not (table.iloc[row, 0] == "" and table.iloc[row, 2] == "" and table.iloc[row, 3] == ""):
                lst_indx = row
            else:
                table.loc[lst_indx, table.columns[1]] = str(table.loc[lst_indx, table.columns[1]]) + str(table.loc[row, table.columns[1]])
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        for row in range(table.shape[0]):
            if re.search("\n-", " ".join(list(table.iloc[row]))) is None:
                table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            else:
                table_row = table.iloc[row]
                for i, cell in enumerate(table_row):
                    if re.search("\n-", cell) is None:
                        table_row[i] = cell.replace('\n', ' ')
        return table


    PKL_NAME = "sp9"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП9 уже существует")
        return True, [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП9
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp9(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]
    

def read_sp10(PATH, f_name):
    
    def add_header_cols_sp10(pages):
        col_header = ""
        pages.insert(4, "Раздел", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)
        for row in range(pages.shape[0]):
            if str(pages.iloc[row, 0]).upper() == str(pages.iloc[row, 0]) and len(pages.iloc[row, 0]) > 2:  # заголовок 1 уровня
                drops.append(row)
                col_header = pages.iloc[row, 0]
            pages.loc[row, pages.columns[4]] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def get_table_documents_sp10(path):
        import re
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = add_header_cols_sp10(table)
        table.drop(0, axis=0, inplace=True)
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            if not (table.iloc[row, 0] == "" and table.iloc[row, 2] == "" and table.iloc[row, 3] == ""):
                lst_indx = row
            else:
                table.loc[lst_indx, table.columns[1]] = str(table.loc[lst_indx, table.columns[1]]) + str(table.loc[row, table.columns[1]])
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        for row in range(table.shape[0]):
            if re.search("\n-", " ".join(list(table.iloc[row]))) is None:
                table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            else:
                table_row = table.iloc[row]
                for i, cell in enumerate(table_row):
                    if re.search("\n-", cell) is None:
                        table_row[i] = cell.replace('\n', ' ')
        return table


    PKL_NAME = "sp10"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП10 уже существует")
        return True, [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП10
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp10(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]
    

def read_sp11(PATH, f_name):
    
    def add_header_cols_sp11(pages):
        col_header = ""
        pages.insert(4, "Раздел", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)
        for row in range(pages.shape[0]):
            if str(pages.iloc[row, 0]).upper() == str(pages.iloc[row, 0]) and len(pages.iloc[row, 0]) > 2:  # заголовок 1 уровня
                drops.append(row)
                col_header = pages.iloc[row, 0]
            pages.loc[row, pages.columns[4]] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def get_table_documents_sp11(path):
        import re
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.drop(table.columns[-1], axis=1, inplace=True)
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = add_header_cols_sp11(table)
        table = table.reset_index(drop=True)
        table.drop(table.columns[0], axis=1, inplace=True)
        table.drop(0, axis=0, inplace=True)
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            if table.iloc[row, 2] is None:
                table.loc[row, table.columns[2]] = table.loc[row-1, table.columns[2]]
            if (table.iloc[row, 0] is not None) and (table.iloc[row, 0] is not None):
                lst_indx = row
            else:
                table.loc[lst_indx, table.columns[2]] = str(table.loc[lst_indx, table.columns[2]]) + str(table.loc[row, table.columns[2]])
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        for row in range(table.shape[0]):
            if re.search("\n-", " ".join(list(table.iloc[row]))) is None:
                table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
        return table


    PKL_NAME = "sp11"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП11 уже существует")
        return True,  [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП11
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp11(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True,  [join(PATH, f'{PKL_NAME}.pkl')]


def read_sp12(PATH, f_name):
    
    def add_header_cols_sp12(pages):
        import re
        from string import digits
        col_header = ""
        pages.insert(3, "Раздел", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)
        upper_letters = [chr(i) for i in range(ord("А"), ord("Я")+1)]
        for row in range(pages.shape[0]):
            if list(pages.iloc[row]).count("") > 2 and pages.iloc[row, 1][0] in upper_letters and len(pages.iloc[row, 1].split()) < 4:  # заголовок 1 уровня
                drops.append(row)
                col_header = pages.iloc[row, 1]
            pages.loc[row, pages.columns[3]] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def get_table_documents_sp12(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = add_header_cols_sp12(table)
        table = table.reset_index(drop=True)
        drops = [0]
        for row in range(table.shape[0]):
            if table.iloc[row, 3] is None:
                table.loc[row, table.columns[3]] = table.loc[row-1, table.columns[3]]
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        drops = [0]
        for row in range(table.shape[0]):
            if table.iloc[row, 0] == "":
                vals_upper = list(table.iloc[row-1])
                vals_lower = list(table.iloc[row])
                vals_lower[-1] = ""
                vals = [vals_upper[i] + vals_lower[i] for i in range(len(vals_lower))]
                table.loc[row-1] = vals
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp12"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП12 уже существует")
        return True, [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП12
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp12(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True,  [join(PATH, f'{PKL_NAME}.pkl')]
    

def read_sp13(PATH, f_name):
    
    def add_header_cols_sp13(pages):
        import re
        from string import digits
        col_header = ""
        pages.insert(4, "Раздел", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)

        for row in range(pages.shape[0]):
            if pages.iloc[row, 0] is not None:
                if list(pages.iloc[row]).count(None) > 2:  # заголовок 1 уровня
                    drops.append(row)
                    if '\n' in pages.iloc[row, 0]:
                        col_header = pages.iloc[row, 0].split('\n')[0]
                    else:
                        col_header = pages.iloc[row, 0]
            pages.at[row, pages.columns[4]] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def get_table_documents_sp13(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = add_header_cols_sp13(table)
        table = table.reset_index(drop=True)
        drops = [0]
        for row in range(table.shape[0]):
            if table.iloc[row, 3] is None:
                table.loc[row, table.columns[3]] = table.loc[row-1, table.columns[3]]
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            if table.iloc[row, 0] == "":
                vals_upper = list(table.iloc[row-1])
                vals_lower = list(table.iloc[row])
                vals_lower[-1] = ""
                vals = [vals_upper[i] + vals_lower[i] for i in range(len(vals_lower))]
                table.loc[row-1] = vals
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp13"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП13 уже существует")
        return True, [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП13
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp13(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]
    

def read_sp14(PATH, f_name):
    
    def add_header_cols_sp14(pages):
        import re
        from string import digits
        col_header = ""
        pages.insert(4, "Раздел", [''] * pages.shape[0])
        drops = []
        pages = pages.reset_index(drop=True)

        for row in range(pages.shape[0]):
            if pages.iloc[row, 0] is not None:
                if list(pages.iloc[row]).count(None) > 2:  # заголовок 1 уровня
                    drops.append(row)
                    if '\n' in pages.iloc[row, 0]:
                        col_header = pages.iloc[row, 0].split('\n')[0]
                    else:
                        col_header = pages.iloc[row, 0]
            pages.at[row, pages.columns[4]] = col_header
        for row in drops:
            pages = pages.drop([row])
        pages = pages.reset_index(drop=True)
        return pages

    def get_table_documents_sp14(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = add_header_cols_sp14(table)
        table = table.reset_index(drop=True)
        drops = [0]
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            if table.iloc[row, 0] == "":
                vals_upper = list(table.iloc[row-1])
                vals_lower = list(table.iloc[row])
                vals = [vals_upper[i] + vals_lower[i] for i in range(len(vals_lower))]
                table.loc[row-1] = vals
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp14"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП14 уже существует")
        return True,  [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП14
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp14(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True,  [join(PATH, f'{PKL_NAME}.pkl')]


def read_sp15(PATH, f_name):
    
    def get_table_documents_sp15_1(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'ПЕРЕЧЕНЬ ДОКУМЕНТОВ'.lower())
        number_end = find_number_page(text, 'Срок хранения'.lower())
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]

        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    def get_table_documents_sp15_2(path):
        table_settings = {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)
        number_begin = find_number_page(text, 'Срок хранения'.lower())
        number_end = len(pdf_res.pages) - 1
        tables = pdf_res.pages[number_begin].find_tables(table_settings)
        table = pd.DataFrame(tables[1].extract())
        for i in range(number_begin+1, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp15"
    if exists(join(PATH, f'{PKL_NAME}_1.pkl')) and exists(join(PATH, f'{PKL_NAME}_2.pkl')):
        print("Файлы СП15 уже существуют")
        return True,  [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]

    # чтение СП15
    ABS_PATH = join(PATH, f_name)
    table1 = get_table_documents_sp15_1(ABS_PATH)
    table2 = get_table_documents_sp15_2(ABS_PATH)
    data1 = table1.reset_index(drop=True)
    data2 = table2.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}_1.pkl', 'wb') as f:
        pkl.dump(data1, f)
        print(f"Файл справочника {PKL_NAME}_1.pkl создан")
    with open(f'{join(PATH, PKL_NAME)}_2.pkl', 'wb') as f:
        pkl.dump(data2, f)
        print(f"Файл справочника {PKL_NAME}_2.pkl создан")
    return True,  [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]
    

def read_sp16(PATH, f_name):
    
    def get_table_documents_sp16_1(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'ПЕРЕЧЕНЬ ДОКУМЕНТОВ'.lower())
        number_end = find_number_page(text, 'Срок хранения'.lower())
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]

        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    def get_table_documents_sp16_2(path):
        table_settings = {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)
        number_begin = find_number_page(text, 'Срок хранения'.lower())
        number_end = len(pdf_res.pages) - 1
        tables = pdf_res.pages[number_begin].find_tables(table_settings)
        table = pd.DataFrame(tables[1].extract())
        for i in range(number_begin+1, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp16"
    if exists(join(PATH, f'{PKL_NAME}_1.pkl')) and exists(join(PATH, f'{PKL_NAME}_2.pkl')):
        print("Файлы СП16 уже существуют")
        return True,  [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]

    # чтение СП16
    ABS_PATH = join(PATH, f_name)
    table1 = get_table_documents_sp16_1(ABS_PATH)
    table2 = get_table_documents_sp16_2(ABS_PATH)
    data1 = table1.reset_index(drop=True)
    data2 = table2.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}_1.pkl', 'wb') as f:
        pkl.dump(data1, f)
        print(f"Файл справочника {PKL_NAME}_1.pkl создан")
    with open(f'{join(PATH, PKL_NAME)}_2.pkl', 'wb') as f:
        pkl.dump(data2, f)
        print(f"Файл справочника {PKL_NAME}_2.pkl создан")
    return True,  [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]
    

def read_sp17(PATH, f_name):
    
    def get_table_documents_sp17_1(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = find_number_page(text, 'Рекомендуется'.lower()) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]

        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table

    def get_table_documents_sp17_2(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Рекомендуется'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]

        table = table.reset_index(drop=True)
        col3_val = ""
        drops = []
        for row in range(table.shape[0]):
            if table.iloc[row, 3] is not None:
                col3_val = table.iloc[row, 3]
            else:
                table.loc[row, table.columns[3]] = col3_val
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp17"
    if exists(join(PATH, f'{PKL_NAME}_1.pkl')) and exists(join(PATH, f'{PKL_NAME}_2.pkl')):
        print("Файлы СП17 уже существуют")
        return True, [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]

    # чтение СП17
    ABS_PATH = join(PATH, f_name)
    table1 = get_table_documents_sp17_1(ABS_PATH)
    table2 = get_table_documents_sp17_2(ABS_PATH)
    data1 = table1.reset_index(drop=True)
    data2 = table2.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}_1.pkl', 'wb') as f:
        pkl.dump(data1, f)
        print(f"Файл справочника {PKL_NAME}_1.pkl создан")
    with open(f'{join(PATH, PKL_NAME)}_2.pkl', 'wb') as f:
        pkl.dump(data2, f)
        print(f"Файл справочника {PKL_NAME}_2.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]
    

def read_sp18(PATH, f_name):
    
    def get_table_documents_sp18(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]

        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp18"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файлы СП18 уже существуют")
        return True, [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП18
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp18(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]


def read_sp19(PATH, f_name):
    
    def get_table_documents_sp19_1(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = find_number_page(text, 'Срок хранения'.lower())
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]

        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    def get_table_documents_sp19_2(path):
        table_settings = {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)
        number_begin = find_number_page(text, 'Срок хранения'.lower())
        tables = pdf_res.pages[number_begin].find_tables(table_settings)
        table = pd.DataFrame(tables[1].extract())
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]
        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp19"
    if exists(join(PATH, f'{PKL_NAME}_1.pkl')) and exists(join(PATH, f'{PKL_NAME}_2.pkl')):
        print("Файлы СП19 уже существуют")
        return True, [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]

    # чтение СП19
    ABS_PATH = join(PATH, f_name)
    table1 = get_table_documents_sp19_1(ABS_PATH)
    table2 = get_table_documents_sp19_2(ABS_PATH)
    data1 = table1.reset_index(drop=True)
    data2 = table2.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}_1.pkl', 'wb') as f:
        pkl.dump(data1, f)
        print(f"Файл справочника {PKL_NAME}_1.pkl создан")
    with open(f'{join(PATH, PKL_NAME)}_2.pkl', 'wb') as f:
        pkl.dump(data2, f)
        print(f"Файл справочника {PKL_NAME}_2.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}_1.pkl'), join(PATH, f'{PKL_NAME}_2.pkl')]


def read_sp20(PATH, f_name):
    
    def get_table_documents_sp20(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]

        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp20"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП20 уже существуют")
        return True, [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП20
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp20(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]


def read_sp21(PATH, f_name):
    
    def get_table_documents_sp21(path):
        table_settings = {
        }

        text, full_text, page_blocks = read_text_documents(path)
        pdf_res = pdfplumber.open(path)

        number_begin = find_number_page(text, 'Перечень документов'.lower())
        number_end = len(pdf_res.pages) - 1
        table = pd.DataFrame()
        for i in range(number_begin, number_end + 1):
            table = pd.concat([table, pd.DataFrame(pdf_res.pages[i].extract_table(table_settings))])
        table.columns = [x.replace('\n', ' ') for x in table.iloc[0]]

        table = table.reset_index(drop=True)
        drops = []
        for row in range(table.shape[0]):
            table.loc[row] = [x.replace('\n', ' ') for x in table.iloc[row]]
            if list(table.iloc[row]) == list(table.columns):
                drops.append(row)
        table = table.drop(drops)
        table = table.reset_index(drop=True)
        return table


    PKL_NAME = "sp21"
    if exists(join(PATH, f'{PKL_NAME}.pkl')):
        print("Файл СП21 уже существует")
        return True, [join(PATH, f'{PKL_NAME}.pkl')]

    # чтение СП21
    ABS_PATH = join(PATH, f_name)
    table = get_table_documents_sp21(ABS_PATH)
    data = table.reset_index(drop=True)
    with open(f'{join(PATH, PKL_NAME)}.pkl', 'wb') as f:
        pkl.dump(data, f)
        print(f"Файл справочника {PKL_NAME}.pkl создан")
    return True, [join(PATH, f'{PKL_NAME}.pkl')]


sp_converter = {1: read_sp1,
                2: read_sp2, 
                3: read_sp3, 
                4: read_sp4,
                5: read_sp5,
                6: read_sp6,
                7: read_sp7, 
                8: read_sp8,
                9: read_sp9,
                10: read_sp10,
                11: read_sp11,
                12: read_sp12,
                13: read_sp13,
                14: read_sp14,
                15: read_sp15,
                16: read_sp16,
                17: read_sp17,
                18: read_sp18,
                19: read_sp19,
                20: read_sp20,
                21: read_sp21}
